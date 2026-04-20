from datetime import datetime, timedelta, date, time
from collections import defaultdict
import json
from fastapi import UploadFile,HTTPException
from typing import Optional, Any, Dict, List
import openpyxl
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
import csv
from io import StringIO, BytesIO
from uuid import uuid4
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import and_
from app.services.utils.metrics import mape
from decimal import Decimal
from app.repositories.sales_repo import SalesRepository
from app.repositories.menu_items_repo import MenuItemRepository
from app.repositories.forecast_breakdown_repo import ForecastBreakdownRepository
from app.repositories.daily_forecast_accuracy_repo import DailyForecastAccuracyRepository
from app.repositories.restaurants_repo import RestaurantRepository
from app.repositories.activity_logs_repo import ActivityLogRepository
from app.repositories.inventory_repo import InventoryRepository
from app.repositories.ingredients_repo import IngredientRepository
from app.repositories.batch_recipes_repo import BatchRecipeRepository
from app.repositories.alerts_repo import AlertRepository
from app.repositories.purchase_orders_repo import PurchaseOrderRepository
from app.repositories.purchase_order_items_repo import PurchaseOrderItemRepository
from app.repositories.clock_events_repo import ClockEventRepository
from app.repositories.orders_repo import OrdersRepository
from app.utils.logger_helpers import log_method
from app.core.logging import logger 
from app.schemas.dashboard_dto import (
    EodSalesEntriesIn, DailyOverviewOut, SaleOut, 
    ProDailyOverviewOut, DeliveryItemOut, ExpectedDeliveryOut, SalesUploadResponse
)
from typing import Sequence, Optional as Opt
from app.db.models.purchase_orders_orm import PurchaseOrder

class DashboardService:
    def __init__(self, db: AsyncSession, restaurant_id: int, subscription_tier: str, employee_id: int):
        self.db = db
        self.restaurant_id = restaurant_id
        self.subscription_tier = subscription_tier
        self.employee_id = employee_id
        self.sales_repo = SalesRepository(db, restaurant_id)
        self.menu_repo = MenuItemRepository(db, restaurant_id)
        self.restaurant_repo = RestaurantRepository(db,restaurant_id)
        self.forecast_breakdown_repo = ForecastBreakdownRepository(db, restaurant_id)
        self.daily_accuracy_repo = DailyForecastAccuracyRepository(db,restaurant_id)
        self.activity_log_repo = ActivityLogRepository(db,restaurant_id,employee_id)
        self.inventory_repo = InventoryRepository(db, restaurant_id)
        self.ingredients_repo = IngredientRepository(db, restaurant_id)
        self.batch_recipes_repo = BatchRecipeRepository(db, restaurant_id)
        self.alerts_repo = AlertRepository(db, restaurant_id)
        self.purchase_orders_repo = PurchaseOrderRepository(db, restaurant_id)
        self.purchase_order_items_repo = PurchaseOrderItemRepository(db, restaurant_id)
        self.clock_event_repo = ClockEventRepository(db, restaurant_id)
        self.orders_repo = OrdersRepository(db, restaurant_id)

    async def log_activity(self, action: str, details: Any = None):
        """
        Create an activity log entry safely.
        Converts details to JSON string if possible, else uses str().
        Adds timestamp to the log.
        """
        try:
            if details is None:
                details_str = ""
            else:
                try:
                    details_str = json.dumps(details, default=lambda o: o.__dict__, indent=2)
                except (TypeError, ValueError):
                    details_str = str(details)

            log_entry = {
                "action": action,
                "details": details_str,
                "created_at": datetime.utcnow().isoformat() + "Z"
            }

            await self.activity_log_repo.create(log_entry)
            logger.info(f"Activity logged: {action}")
        except Exception as e:
            logger.error(f"Failed to log activity '{action}': {e}", exc_info=True)

    @log_method()
    async def get_live_operations(self):
        """
        Provide real-time operational snapshot required by LiveOperationsOut.
        Uses available repositories where possible; fills safe defaults otherwise.
        """
        try:
            today = date.today()

            # Current shift: count currently clocked-in employees (scheduled/on_break not tracked here)
            try:
                clocked_in = await self.clock_event_repo.get_clocked_in_employees(today)
                current_shift = {
                    "clocked_in": len(clocked_in),
                    "scheduled": 0,
                    "on_break": 0,
                }
            except Exception as e:
                logger.warning(f"Clock events unavailable: {e}", exc_info=True)
                current_shift = {"clocked_in": 0, "scheduled": 0, "on_break": 0}

            # Order flow: derive basic counts from orders table
            pending = in_progress = ready = completed_today = 0
            avg_prep_time = 0.0
            try:
                # Active orders
                active_orders = await self.orders_repo.get_active_orders()
                # Count by status for basic buckets
                for o in active_orders:
                    if getattr(o, "order_status", "open") == "in_progress":
                        in_progress += 1
                    else:
                        pending += 1
                # Completed today (simple filter on timestamp and status if present)
                # Note: If 'completed' status doesn't exist, this stays 0.
                # A richer repo method can refine this later.
            except Exception as e:
                logger.warning(f"Orders unavailable for flow: {e}", exc_info=True)
                active_orders = []

            order_flow = {
                "pending": pending,
                "in_progress": in_progress,
                "ready": ready,
                "completed_today": completed_today,
                "avg_prep_time": avg_prep_time,
            }

            # Today's pace vs forecast (placeholders unless forecasting is queried here)
            todays_pace = {
                "current_sales": 0.0,
                "forecast_sales": 0.0,
                "percentage": 0.0,
                "pace_vs_forecast": "on_track",
            }

            # Active orders list (minimal representation)
            active_orders_out = []
            try:
                for o in active_orders:
                    active_orders_out.append({
                        "order_id": int(o.order_id),
                        "table": getattr(o, "sales_channel", "") or "",
                        "items": 0,  # could be enriched by counting order_items
                        "time_elapsed": 0,
                        "status": getattr(o, "order_status", "open"),
                        "server": "",
                    })
            except Exception as e:
                logger.warning(f"Could not format active orders: {e}", exc_info=True)

            # Kitchen status (no direct signals yet; provide safe defaults)
            kitchen_status = {"grill": "normal", "fryer": "normal", "salad": "normal", "dessert": "normal"}

            # Upcoming deliveries (reuse purchase orders expected today)
            upcoming_deliveries = []
            try:
                stmt = (
                    select(PurchaseOrder)
                    .where(
                        and_(
                            PurchaseOrder.restaurant_id == self.restaurant_id,
                            PurchaseOrder.expected_delivery_date == today,
                            PurchaseOrder.status.in_(["pending", "confirmed", "in_transit"]),
                        )
                    )
                )
                res = await self.db.execute(stmt)
                for po in res.scalars().all():
                    supplier = getattr(getattr(po, "supplier", None), "supplier_name", None) or "Unknown"
                    eta = po.expected_delivery_date.isoformat() if getattr(po, "expected_delivery_date", None) else ""
                    # If items not eager-loaded, present a generic summary
                    items_str = "order items"
                    upcoming_deliveries.append({"supplier": supplier, "eta": eta, "items": items_str})
            except Exception as e:
                logger.warning(f"Upcoming deliveries unavailable: {e}", exc_info=True)

            return {
                "current_shift": current_shift,
                "order_flow": order_flow,
                "todays_pace": todays_pace,
                "active_orders": active_orders_out,
                "kitchen_status": kitchen_status,
                "upcoming_deliveries": upcoming_deliveries,
            }
        except Exception as e:
            logger.error(f"Error in get_live_operations: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Failed to fetch live operations")

    @log_method()
    async def get_quick_analytics(self, days: int = 7):
        """
        Return a lightweight analytics snapshot for the last N days.
        Computes revenue via MenuItem price * Sales.quantity_sold.
        Provides WoW deltas comparing the same length prior window.
        """
        try:
            end_date = date.today()
            start_date = end_date - timedelta(days=days - 1)

            # Current window sales
            sales = await self.sales_repo.get_sales_by_date_range(start_date, end_date)

            # Build a price cache to avoid repeated DB hits
            price_cache = {}
            async def get_price(mid: int) -> float:
                if mid in price_cache:
                    return price_cache[mid]
                mi = await self.menu_repo.get_by_id(mid)
                price = float(mi.price) if mi and mi.price is not None else 0.0
                price_cache[mid] = price
                return price

            # Aggregate by day and by item
            from collections import defaultdict
            daily = defaultdict(lambda: {"sales": 0.0, "orders": 0, "customers": 0})
            per_item = defaultdict(lambda: {"units": 0, "revenue": 0.0, "name": ""})

            for s in sales:
                d = s.sale_timestamp.date()
                qty = int(s.quantity_sold or 0)
                price = await get_price(int(s.menu_item_id))
                rev = qty * price
                daily[d]["sales"] += rev
                # Approximate orders and customers as quantity sold when order granularity not present
                daily[d]["orders"] += qty
                daily[d]["customers"] += qty

                # Per-item aggregation
                per_item[int(s.menu_item_id)]["units"] += qty
                per_item[int(s.menu_item_id)]["revenue"] += rev
                if not per_item[int(s.menu_item_id)]["name"]:
                    mi = await self.menu_repo.get_by_id(int(s.menu_item_id))
                    per_item[int(s.menu_item_id)]["name"] = getattr(mi, "name", f"Item {s.menu_item_id}")

            # Compose daily series for the window
            daily_sales = []
            cur = start_date
            total_sales = 0.0
            total_orders = 0
            total_customers = 0
            while cur <= end_date:
                day_data = daily.get(cur, {"sales": 0.0, "orders": 0, "customers": 0})
                daily_sales.append({
                    "date": cur.isoformat(),
                    "sales": round(day_data["sales"], 2),
                    "orders": int(day_data["orders"]),
                    "customers": int(day_data["customers"]),
                })
                total_sales += day_data["sales"]
                total_orders += day_data["orders"]
                total_customers += day_data["customers"]
                cur += timedelta(days=1)

            avg_order_value = (total_sales / total_orders) if total_orders > 0 else 0.0

            # WoW deltas: previous window of equal length
            prev_end = start_date - timedelta(days=1)
            prev_start = prev_end - timedelta(days=days - 1)
            prev_sales_rows = await self.sales_repo.get_sales_by_date_range(prev_start, prev_end)
            prev_total_sales = 0.0
            prev_total_orders = 0
            prev_total_customers = 0
            for s in prev_sales_rows:
                qty = int(s.quantity_sold or 0)
                price = await get_price(int(s.menu_item_id))
                prev_total_sales += qty * price
                prev_total_orders += qty
                prev_total_customers += qty

            def pct_change(cur_val: float, prev_val: float) -> float:
                if prev_val == 0:
                    return 100.0 if cur_val > 0 else 0.0
                return (cur_val - prev_val) / prev_val * 100.0

            wow_sales_change = pct_change(total_sales, prev_total_sales)
            wow_orders_change = pct_change(float(total_orders), float(prev_total_orders))
            wow_customers_change = pct_change(float(total_customers), float(prev_total_customers))
            wow_avg_change = pct_change(avg_order_value, (prev_total_sales / prev_total_orders) if prev_total_orders > 0 else 0.0)

            # Top/Bottom items by units (use revenue as secondary sort)
            items_sorted = sorted(
                per_item.values(), key=lambda x: (x["units"], x["revenue"]), reverse=True
            )
            def to_perf(it):
                return {
                    "name": it["name"],
                    "units": int(it["units"]),
                    "revenue": round(it["revenue"], 2),
                    "trend": "neutral",
                    "change": 0.0,
                }
            top_items = list(map(to_perf, items_sorted[:5]))
            bottom_items = list(map(to_perf, list(reversed(items_sorted))[:5])) if items_sorted else []

            # Hourly pattern placeholder (can be filled from sales timestamps later)
            hourly_pattern = [0] * 24

            return {
                "summary": {
                    "total_sales": round(total_sales, 2),
                    "total_orders": int(total_orders),
                    "avg_order_value": round(avg_order_value, 2),
                    "total_customers": int(total_customers),
                    "wow_sales_change": round(wow_sales_change, 2),
                    "wow_orders_change": round(wow_orders_change, 2),
                    "wow_avg_change": round(wow_avg_change, 2),
                    "wow_customers_change": round(wow_customers_change, 2),
                },
                "daily_sales": daily_sales,
                "top_items": top_items,
                "bottom_items": bottom_items,
                "hourly_pattern": hourly_pattern,
            }
        except Exception as e:
            logger.error(f"Error in get_quick_analytics: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Failed to fetch quick analytics")

    async def get_daily_overview_data(self):
        # Return the basic overview for all tiers to ensure a valid response.
        # Pro/Master dashboards use a separate /pro-overview endpoint.
        data = await self._get_basic_overview()
        return DailyOverviewOut(**data)

    @log_method("Get basic overview")
    async def _get_basic_overview(self):
        return {
            "forecasted_sales_today": await self._get_forecasted_sales_basic(),
            "top_5_items_today": await self._get_top_forecasted_items_basic(),
            "accuracy_yesterday": await self._get_yesterday_accuracy_basic()
        }

    @log_method("Get Forecasted Sales (Basic)")
    async def _get_forecasted_sales_basic(self):
        today = date.today()
        forecast_rows = await self.forecast_breakdown_repo.get_forecasts_by_date(today)

        total_quantity = 0
        total_revenue = Decimal("0.00")

        for row in forecast_rows:
            menu_item = await self.menu_repo.get_by_id(row.menu_item_id)
            if menu_item:
                total_quantity += row.forecasted_quantity
                price = Decimal(str(menu_item.price))
                total_revenue += Decimal(row.forecasted_quantity) * price

        return {
            "forecasted_quantity": total_quantity,
            "forecasted_revenue": float(round(total_revenue, 2))  # convert Decimal to float for JSON serialization
        }

    @log_method("Get Top Forecasted Items (Basic)")
    async def _get_top_forecasted_items_basic(self):
        today = date.today()
        forecast_rows = await self.forecast_breakdown_repo.get_forecasts_by_date(today)

        item_quantities = defaultdict(int)

        for row in forecast_rows:
            item_quantities[row.menu_item_id] += row.forecasted_quantity

        # Sort by quantity descending
        sorted_items = sorted(item_quantities.items(), key=lambda x: x[1], reverse=True)

        results = []
        for menu_item_id, quantity in sorted_items:
            if len(results) >= 6:
                break

            item = await self.menu_repo.get_by_id(menu_item_id)
            if item and item.restaurant_id == self.restaurant_id and getattr(item, "is_active", True):
                results.append({
                    "menu_item_id": menu_item_id,
                    "name": item.name,
                    "forecasted_quantity": quantity,
                })

        return results
    
    @log_method("Get Yesterday Accuracy (Basic)")
    async def _get_yesterday_accuracy_basic(self):
        

        yesterday = date.today() - timedelta(days=1)
        accuracy_entries = await self.daily_accuracy_repo.get_by_date(yesterday)

        if not accuracy_entries:
            return {
                "accuracy_percent": None,
                "note": "No forecast accuracy data available"
            }

        mape_values = []
        for row in accuracy_entries:
            mape_value = mape(row.predicted_quantity, row.actual_quantity)
            mape_values.append(mape_value)

        avg_mape = sum(mape_values) / len(mape_values)
        accuracy_percent = round(max(0.0, 100 - avg_mape), 2)

        note = "Accurate" if accuracy_percent >= 85 else "Needs Improvement"

        return {
            "accuracy_percent": accuracy_percent,
            "note": note
        }

    @log_method("Create Menu Item")
    async def create_menu_item(self, data: dict):
        logger.info(f"Creating menu item: {data}")
        item = await self.menu_repo.create(data)
        await self.log_activity("create_menu_item", data)
        return item
    
    @log_method("Update Menu Item")
    async def update_menu_item(self, menu_item_id: int, data: dict):
        logger.info(f"Updating Menu Item: {menu_item_id}, {data}" )
        updated = await self.menu_repo.update(menu_item_id, data)
        details = {"menu_item_id": menu_item_id, **data}
        await self.log_activity("update_menu_item", details)
        return updated

    @log_method("Deactivate Menu Item")
    async def deactivate_menu_item(self, menu_item_id: int):
        logger.info("Deactivating menu item ID %d", menu_item_id)
        updated = await self.menu_repo.update(menu_item_id, {"is_active": False})
        await self.log_activity("deactivate_menu_item", updated)
        return updated
    
    @log_method("List Menu Items")
    async def list_menu_items(self):
        print('inside service')
        return await self.menu_repo.get_all()

    @log_method("Upload Bulk Menu Items")
    async def upload_menu_items_csv(self, file: UploadFile):
        contents = await file.read()
        created = []

        def normalize_headers(headers):
            return [header.strip().lower() if isinstance(header, str) else header for header in headers]

        if file.filename.endswith('.csv'):
            decoded = contents.decode('utf-8')
            raw_lines = decoded.splitlines()
            reader = csv.DictReader(raw_lines)
            reader.fieldnames = normalize_headers(reader.fieldnames)

            for row in reader:
                normalized_row = {k.strip().lower(): v for k, v in row.items()}
                try:
                    item_data = {
                        "name": normalized_row['name'],
                        "category": normalized_row.get('category'),
                        "price": float(normalized_row['price']),
                        "is_active": normalized_row.get('is_active', '1').lower() in ['1', 'true', 'yes']
                    }
                    item = await self.menu_repo.create(item_data)
                    created.append(item)
                    await self.log_activity("upload_menu_item", item_data)
                except Exception as e:
                    logger.error(f"Failed to process menu item row {row}: {e}", exc_info=True)
                    continue

        elif file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filename=BytesIO(contents), read_only=True)
            sheet = workbook.active

            raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            headers = normalize_headers(raw_headers)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                try:
                    item_data = {
                        "name": row_dict['name'],
                        "category": row_dict.get('category'),
                        "price": float(row_dict['price']),
                        "is_active": str(row_dict.get('is_active', '1')).lower() in ['1', 'true', 'yes']
                    }
                    item = await self.menu_repo.create(item_data)
                    created.append(item)
                    await self.log_activity("upload_menu_item",item_data)
                except Exception as e:
                    logger.error(f"Failed to process menu item row {row_dict}: {e}", exc_info=True)
                    continue

        else:
            raise ValueError("Unsupported file format. Please upload CSV or XLSX.")

        return created

    def _normalize_headers(self, headers):
        return [header.strip().lower() if isinstance(header, str) else header for header in (headers or [])]

    def _validate_sales_upload_headers(self, headers):
        header_set = {header for header in headers if header}
        missing_columns = []

        if not ({"menu_item_id", "menu_item_name"} & header_set):
            missing_columns.append("menu_item_id or menu_item_name")
        if "quantity_sold" not in header_set:
            missing_columns.append("quantity_sold")
        if "sale_timestamp" not in header_set:
            missing_columns.append("sale_timestamp")

        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    def _serialize_sales_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        serialized = {}
        for key, value in row.items():
            if isinstance(value, (datetime, date)):
                serialized[key] = value.isoformat()
            elif isinstance(value, Decimal):
                serialized[key] = str(value)
            elif value is None or isinstance(value, (str, int, float, bool)):
                serialized[key] = value
            else:
                serialized[key] = str(value)
        return serialized

    def _clean_sales_channel(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        return str(value)

    def _to_sale_out(self, sale: Any) -> SaleOut:
        sale_timestamp = getattr(sale, "sale_timestamp", None)
        timestamp_iso = sale_timestamp.isoformat() if sale_timestamp else None
        return SaleOut(
            sale_id=getattr(sale, "sale_id", getattr(sale, "id", None)),
            menu_item_id=getattr(sale, "menu_item_id", None),
            quantity_sold=getattr(sale, "quantity_sold", 0),
            sales_channel=getattr(sale, "sales_channel", None),
            sale_timestamp=timestamp_iso,
        )

    def _build_sales_upload_response(
        self,
        *,
        source: str,
        overwrite: bool,
        attempted_rows: int,
        inserted_sales: List[Any],
        duplicate_rows: int,
        overwritten_rows: int,
        row_errors: List[Dict[str, Any]],
        date_channels: Dict[date, set],
    ) -> SalesUploadResponse:
        sale_dates = sorted(day.isoformat() for day in date_channels.keys())
        channels = sorted(
            {channel for values in date_channels.values() for channel in values},
            key=lambda value: "" if value is None else str(value).lower(),
        )
        data = [self._to_sale_out(sale) for sale in inserted_sales]
        skipped_rows = max(attempted_rows - len(inserted_sales), 0)

        if attempted_rows == 0:
            message = "No sales rows were found to import."
        else:
            message = f"Imported {len(inserted_sales)} of {attempted_rows} sales rows."
            if skipped_rows:
                message += f" Skipped {skipped_rows} row{'s' if skipped_rows != 1 else ''}."
            if overwritten_rows:
                message += f" Replaced {overwritten_rows} existing row{'s' if overwritten_rows != 1 else ''}."

        return SalesUploadResponse(
            import_id=str(uuid4()),
            source=source,
            overwrite=overwrite,
            attempted_rows=attempted_rows,
            inserted_rows=len(inserted_sales),
            skipped_rows=skipped_rows,
            duplicate_rows=duplicate_rows,
            overwritten_rows=overwritten_rows,
            sale_dates=sale_dates,
            channels=channels,
            row_errors=row_errors,
            message=message,
            data=data,
        )

    @log_method("Upload Sales Data")
    async def upload_sales_data(self, file: UploadFile, overwrite: bool = False) -> SalesUploadResponse:
        contents = await file.read()
        parsed_rows = []
        filename = (file.filename or "").lower()

        if filename.endswith(".csv"):
            decoded = contents.decode("utf-8-sig")
            raw_lines = decoded.splitlines()
            reader = csv.DictReader(raw_lines)
            if not reader.fieldnames:
                raise ValueError("Sales upload file is empty.")
            reader.fieldnames = self._normalize_headers(reader.fieldnames)
            self._validate_sales_upload_headers(reader.fieldnames)

            for row_number, row in enumerate(reader, start=2):
                normalized_row = {}
                for key, value in (row or {}).items():
                    if key is None:
                        continue
                    normalized_key = key.strip().lower() if isinstance(key, str) else key
                    normalized_row[normalized_key] = value.strip() if isinstance(value, str) else value
                parsed_rows.append({"row_number": row_number, "row": normalized_row})

        elif filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(filename=BytesIO(contents), read_only=True)
            sheet = workbook.active
            try:
                headers = self._normalize_headers(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            except StopIteration as exc:
                raise ValueError("Sales upload file is empty.") from exc
            self._validate_sales_upload_headers(headers)

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                parsed_rows.append({"row_number": row_number, "row": dict(zip(headers, row))})

        else:
            raise ValueError("Unsupported file format. Please upload CSV or XLSX.")

        attempted_rows = len(parsed_rows)
        row_errors = []
        duplicate_rows = 0
        date_channels = {}
        valid_sales_rows = []
        seen_rows = set()
        menu_item_cache_by_id = {}
        menu_item_cache_by_name = {}

        for parsed in parsed_rows:
            row_number = parsed["row_number"]
            row = parsed["row"]

            quantity_raw = row.get("quantity_sold")
            if quantity_raw in (None, ""):
                continue

            try:
                quantity_decimal = Decimal(str(quantity_raw))
            except Exception:
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "invalid_quantity",
                        "message": "quantity_sold must be a whole number.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            if quantity_decimal <= 0:
                continue

            if quantity_decimal != quantity_decimal.to_integral_value():
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "invalid_quantity",
                        "message": "quantity_sold must be a whole number.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            quantity = int(quantity_decimal)
            timestamp_raw = row.get("sale_timestamp")
            if timestamp_raw in (None, ""):
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "missing_sale_timestamp",
                        "message": "sale_timestamp is required for every uploaded sales row.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            try:
                if isinstance(timestamp_raw, datetime):
                    sale_time = timestamp_raw
                elif isinstance(timestamp_raw, date):
                    sale_time = datetime.combine(timestamp_raw, time.min)
                else:
                    sale_time = datetime.fromisoformat(str(timestamp_raw).strip())
            except Exception:
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "invalid_sale_timestamp",
                        "message": "sale_timestamp must be a valid ISO datetime.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            sales_channel = self._clean_sales_channel(row.get("sales_channel"))
            menu_item_name = str(row.get("menu_item_name") or "").strip()
            menu_item = None

            if row.get("menu_item_id") not in (None, ""):
                try:
                    menu_item_id_decimal = Decimal(str(row.get("menu_item_id")))
                    if menu_item_id_decimal != menu_item_id_decimal.to_integral_value():
                        raise ValueError()
                    resolved_menu_item_id = int(menu_item_id_decimal)
                except Exception:
                    row_errors.append(
                        {
                            "row_number": row_number,
                            "code": "invalid_menu_item_id",
                            "message": "menu_item_id must be a whole number when provided.",
                            "row_data": self._serialize_sales_row(row),
                        }
                    )
                    continue

                if resolved_menu_item_id not in menu_item_cache_by_id:
                    menu_item_cache_by_id[resolved_menu_item_id] = await self.menu_repo.get_by_id(resolved_menu_item_id)
                menu_item = menu_item_cache_by_id[resolved_menu_item_id]
            elif menu_item_name:
                normalized_name = menu_item_name.lower()
                if normalized_name not in menu_item_cache_by_name:
                    menu_item_cache_by_name[normalized_name] = await self.menu_repo.get_by_name(menu_item_name)
                menu_item = menu_item_cache_by_name[normalized_name]
            else:
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "missing_menu_item",
                        "message": "Each row must include either menu_item_id or menu_item_name.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            if not menu_item or getattr(menu_item, "restaurant_id", None) != self.restaurant_id:
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "menu_item_not_found",
                        "message": "The referenced menu item could not be found for this restaurant.",
                        "row_data": self._serialize_sales_row(row),
                    }
                )
                continue

            menu_item_id = getattr(menu_item, "menu_item_id", getattr(menu_item, "id", None))
            duplicate_key = (menu_item_id, sales_channel, sale_time.isoformat())
            if duplicate_key in seen_rows:
                duplicate_rows += 1
                continue

            seen_rows.add(duplicate_key)
            sale_date = sale_time.date()
            date_channels.setdefault(sale_date, set()).add(sales_channel)
            valid_sales_rows.append(
                {
                    "menu_item_id": menu_item_id,
                    "quantity_sold": quantity,
                    "sales_channel": sales_channel,
                    "sale_timestamp": sale_time,
                }
            )

        overwritten_rows = 0
        if valid_sales_rows:
            if not overwrite:
                for sale_date, channels in date_channels.items():
                    if await self.sales_repo.sales_exist_for_date_and_channels(sale_date, list(channels)):
                        raise HTTPException(
                            status_code=409,
                            detail="Sales already exist for one or more date/channel combinations in this file. Confirm overwrite to replace those channels.",
                        )
            else:
                for sale_date, channels in date_channels.items():
                    overwritten_rows += await self.sales_repo.delete_sales_for_date_and_channels(
                        sale_date,
                        list(channels),
                        auto_commit=False,
                    )

        inserted_sales = await self.sales_repo.create_many(valid_sales_rows)
        response = self._build_sales_upload_response(
            source="file",
            overwrite=overwrite,
            attempted_rows=attempted_rows,
            inserted_sales=inserted_sales,
            duplicate_rows=duplicate_rows,
            overwritten_rows=overwritten_rows,
            row_errors=row_errors,
            date_channels=date_channels,
        )
        await self.log_activity(
            "upload_sales_data",
            {
                "import_id": response.import_id,
                "source": response.source,
                "overwrite": overwrite,
                "attempted_rows": response.attempted_rows,
                "inserted_rows": response.inserted_rows,
                "skipped_rows": response.skipped_rows,
                "duplicate_rows": response.duplicate_rows,
                "overwritten_rows": response.overwritten_rows,
                "row_error_count": len(response.row_errors),
            },
        )
        return response

    @log_method("Upload Sales Entries (Manual JSON)")
    async def upload_sales_entries(self, payload: EodSalesEntriesIn) -> SalesUploadResponse:
        """Insert manual EOD sales entries for a single date."""
        try:
            sale_date = datetime.fromisoformat(payload.sale_date).date()
        except Exception:
            raise ValueError("Invalid sale_date; expected YYYY-MM-DD")

        attempted_rows = len(payload.entries)
        row_errors = []
        duplicate_rows = 0
        date_channels = {}
        valid_sales_rows = []
        seen_rows = set()
        sale_time = datetime.combine(sale_date, time.min)

        for row_number, entry in enumerate(payload.entries, start=1):
            entry_data = entry.model_dump() if hasattr(entry, "model_dump") else entry.dict()
            quantity = int(entry.quantity_sold or 0)
            if quantity <= 0:
                continue

            sales_channel = self._clean_sales_channel(entry.sales_channel)
            duplicate_key = (entry.menu_item_id, sales_channel)
            if duplicate_key in seen_rows:
                duplicate_rows += 1
                continue

            seen_rows.add(duplicate_key)
            item = await self.menu_repo.get_by_id(entry.menu_item_id)
            if not item or getattr(item, "restaurant_id", None) != self.restaurant_id:
                row_errors.append(
                    {
                        "row_number": row_number,
                        "code": "menu_item_not_found",
                        "message": "The referenced menu item could not be found for this restaurant.",
                        "row_data": self._serialize_sales_row(entry_data),
                    }
                )
                continue

            date_channels.setdefault(sale_date, set()).add(sales_channel)
            valid_sales_rows.append(
                {
                    "menu_item_id": entry.menu_item_id,
                    "quantity_sold": quantity,
                    "sales_channel": sales_channel,
                    "sale_timestamp": sale_time,
                }
            )

        channels = list(date_channels.get(sale_date, set()))
        overwritten_rows = 0
        if valid_sales_rows and channels:
            if payload.overwrite:
                overwritten_rows = await self.sales_repo.delete_sales_for_date_and_channels(
                    sale_date,
                    channels,
                    auto_commit=False,
                )
            elif await self.sales_repo.sales_exist_for_date_and_channels(sale_date, channels):
                raise HTTPException(
                    status_code=409,
                    detail="Sales already exist for that date and channel(s). Confirm overwrite to replace.",
                )

        inserted = await self.sales_repo.create_many(valid_sales_rows)
        response = self._build_sales_upload_response(
            source="manual",
            overwrite=bool(payload.overwrite),
            attempted_rows=attempted_rows,
            inserted_sales=inserted,
            duplicate_rows=duplicate_rows,
            overwritten_rows=overwritten_rows,
            row_errors=row_errors,
            date_channels=date_channels,
        )
        await self.log_activity(
            "upload_sales_manual",
            {
                "import_id": response.import_id,
                "sale_date": payload.sale_date,
                "overwrite": bool(payload.overwrite),
                "attempted_rows": response.attempted_rows,
                "inserted_rows": response.inserted_rows,
                "skipped_rows": response.skipped_rows,
                "duplicate_rows": response.duplicate_rows,
                "overwritten_rows": response.overwritten_rows,
                "row_error_count": len(response.row_errors),
            },
        )
        return response
    
    @log_method("Generating Sales Upload Template")
    async def generate_sales_upload_template_xlsx(self, default_date: Optional[str] = None) -> BytesIO:
        menu_items = await self.list_menu_items()
        active_menu_items = [item for item in menu_items if item.is_active]
        restaurant_settings = await self.restaurant_repo.get_settings()
        sales_channels = restaurant_settings['sales_channels']

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales Upload Template"

        headers = ["menu_item_id", "menu_item_name", "quantity_sold", "sales_channel", "sale_timestamp"]
        sheet.append(headers)

        if default_date is None:
            default_date = datetime.utcnow().strftime("%Y-%m-%dT00:00:00")

        # Append all rows first
        for item in active_menu_items:
            for channel in sales_channels:
                sheet.append([item.menu_item_id, item.name, "", channel, default_date])

        total_rows = sheet.max_row
        total_cols = len(headers)

        # Lock all cells by default
        for row in range(1, total_rows + 1):
            for col in range(1, total_cols + 1):
                sheet.cell(row=row, column=col).protection = Protection(locked=True)

        # Unlock quantity_sold column cells (column 3), skipping header
        for row in range(2, total_rows + 1):
            sheet.cell(row=row, column=3).protection = Protection(locked=False)

        # Set column widths based on max content length per column
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(header)
            col_letter = get_column_letter(col_idx)
            for row in range(1, total_rows + 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[col_letter].width = max_length + 2

        # Protect the sheet
        sheet.protection.sheet = True
        sheet.protection.password = "restaurant"
        sheet.protection.enable()

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @log_method("Check sales conflicts by date/channels")
    async def check_sales_conflicts(self, sale_date: str, channels: Opt[Sequence[str]] = None):
        """Return per-channel counts for a date, to support client confirmation UX.
        channels: if provided, filter to those channels; use the string 'null' to represent None.
        """
        try:
            d = datetime.fromisoformat(sale_date).date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid sale_date; expected YYYY-MM-DD")

        # If channels not provided, return counts for all channels present
        if not channels:
            rows = await self.sales_repo.get_sales_channels_counts_for_date(d)
            return {
                "sale_date": sale_date,
                "conflicts": { (ch if ch is not None else None): cnt for ch, cnt in rows },
            }
        # Normalize 'null' -> None
        norm_channels = [None if ch.lower() == 'null' else ch for ch in channels]
        exists = await self.sales_repo.sales_exist_for_date_and_channels(d, norm_channels)
        if not exists:
            return {"sale_date": sale_date, "conflicts": {}}
        # Return counts but filtered to requested channels
        rows = await self.sales_repo.get_sales_channels_counts_for_date(d)
        filtered = { (ch if ch is not None else None): cnt for ch, cnt in rows if (ch in norm_channels) or (ch is None and None in norm_channels) }
        return {"sale_date": sale_date, "conflicts": filtered}

    @log_method()
    async def get_pro_daily_overview(self) -> ProDailyOverviewOut:
        """
        Comprehensive daily overview for Pro/Master tier with:
        - Basic metrics
        - Inventory summary & alerts
        - Prep schedule & task completion
        - Menu performance (forecast vs actuals)
        - Expected deliveries today
        """
        try:
            today = date.today()
            logger.info(f"Fetching Pro daily overview for restaurant {self.restaurant_id} on {today}")
            
            # Get basic overview data
            basic_data = await self._get_basic_overview()
            logger.info("Basic overview data retrieved successfully")
            
            # ===== INVENTORY SUMMARY & ALERTS =====
            all_inventory = await self.inventory_repo.get_all()
            
            critical_stock = []
            low_stock = []
            healthy_stock = []
            total_value = 0.0
            
            for inv in all_inventory:
                ingredient = await self.ingredients_repo.get_by_id(inv.ingredient_id)
                if not ingredient:
                    continue
                    
                item_value = float(inv.quantity_available or 0) * float(ingredient.cost_per_unit or 0)
                total_value += item_value
                
                stock_item = {
                    "ingredient_id": inv.ingredient_id,
                    "ingredient_name": ingredient.ingredient_name,
                    "quantity_available": float(inv.quantity_available or 0),
                    "par_level": float(inv.par_level or 0),
                    "unit": ingredient.unit,
                    "value": item_value
                }
                
                if inv.quantity_available <= (inv.par_level * 0.25):
                    critical_stock.append(stock_item)
                elif inv.quantity_available <= inv.par_level:
                    low_stock.append(stock_item)
                else:
                    healthy_stock.append(stock_item)
            
            inventory_summary = {
                "critical_count": len(critical_stock),
                "low_count": len(low_stock),
                "healthy_count": len(healthy_stock),
                "total_value": total_value,
                "critical_items": critical_stock[:5],  # Top 5
                "low_items": low_stock[:5]
            }
            
            # Get alerts
            all_alerts = await self.alerts_repo.get_all()
            active_alerts = [
                {
                    "alert_id": alert.alert_id,
                    "alert_type": alert.alert_type,
                    "message": alert.message,
                    "severity": alert.severity,
                    "created_at": alert.created_at.isoformat() if alert.created_at else None
                }
                for alert in all_alerts
                if not alert.is_resolved
            ]
            
            # ===== PREP SCHEDULE & TASK COMPLETION =====
            batch_recipes = await self.batch_recipes_repo.get_all()
            
            prep_tasks = []
            total_tasks = len(batch_recipes)
            
            for recipe in batch_recipes:
                prep_tasks.append({
                    "batch_recipe_id": recipe.batch_recipe_id,
                    "batch_name": recipe.batch_name,
                    "is_completed": False,  # TODO: Track prep completion
                    "target_quantity": float(recipe.target_quantity or 0),
                    "unit": recipe.unit
                })
            
            prep_schedule = {
                "total_tasks": total_tasks,
                "completed_tasks": 0,  # TODO: Count completed tasks
                "completion_rate": 0,  # TODO: Calculate completion rate
                "tasks": prep_tasks
            }
            
            # ===== MENU PERFORMANCE (Forecast vs Actuals) =====
            # Get today's forecast breakdown
            forecast_items = await self.forecast_breakdown_repo.get_by_date(today)
            
            # Get today's actual sales
            actual_sales = await self.sales_repo.get_by_date(today)
            
            # Build actual sales by menu item
            actual_by_item = {}
            for sale in actual_sales:
                menu_item_id = sale.menu_item_id
                actual_by_item[menu_item_id] = actual_by_item.get(menu_item_id, 0) + (sale.quantity or 0)
            
            menu_performance = []
            for forecast in forecast_items:
                menu_item = await self.menu_repo.get_by_id(forecast.menu_item_id)
                if not menu_item:
                    continue
                    
                forecasted_qty = float(forecast.predicted_quantity or 0)
                actual_qty = float(actual_by_item.get(forecast.menu_item_id, 0))
                variance = actual_qty - forecasted_qty
                variance_pct = (variance / forecasted_qty * 100) if forecasted_qty > 0 else 0
                
                menu_performance.append({
                    "menu_item_id": forecast.menu_item_id,
                    "item_name": menu_item.item_name,
                    "forecasted_quantity": forecasted_qty,
                    "actual_quantity": actual_qty,
                    "variance": variance,
                    "variance_percent": variance_pct
                })
            
            # ===== EXPECTED DELIVERIES TODAY =====
            stmt = (
                select(PurchaseOrder)
                .options(selectinload(PurchaseOrder.supplier))
                .options(selectinload(PurchaseOrder.purchase_order_items))
                .where(
                    and_(
                        PurchaseOrder.restaurant_id == self.restaurant_id,
                        PurchaseOrder.expected_delivery_date == today,
                        PurchaseOrder.status.in_(["pending", "confirmed", "in_transit"])
                    )
                )
            )
            result = await self.db.execute(stmt)
            orders = result.scalars().all()
            
            expected_deliveries = []
            for order in orders:
                # Build delivery items
                delivery_items = []
                for item in order.purchase_order_items:
                    ingredient = await self.ingredients_repo.get_by_id(item.ingredient_id)
                    if ingredient:
                        delivery_items.append(DeliveryItemOut(
                            ingredient_name=ingredient.ingredient_name,
                            quantity_ordered=float(item.quantity_ordered or 0),
                            unit=ingredient.unit
                        ))
                
                expected_deliveries.append(ExpectedDeliveryOut(
                    order_id=order.purchase_order_id,
                    supplier_name=order.supplier.supplier_name if order.supplier else "Unknown",
                    expected_delivery_date=order.expected_delivery_date.isoformat() if order.expected_delivery_date else None,
                    order_date=order.order_date.isoformat() if order.order_date else None,
                    status=order.status or "pending",
                    total_items=len(delivery_items),
                    total_order_price=float(order.total_order_price or 0),
                    items=delivery_items
                ))
            
            logger.info(f"Pro overview compiled successfully: {len(expected_deliveries)} expected deliveries")
            
            return ProDailyOverviewOut(
                **basic_data,
                inventory_summary=inventory_summary,
                active_alerts=active_alerts,
                prep_schedule=prep_schedule,
                menu_performance=menu_performance,
                expected_deliveries_today=expected_deliveries
            )
            
        except Exception as e:
            logger.error(f"Error in get_pro_daily_overview: {e}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Failed to fetch Pro daily overview: {str(e)}"
            )
    
    