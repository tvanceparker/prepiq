from datetime import datetime, timedelta, date
from collections import defaultdict
import json
from fastapi import UploadFile,HTTPException
from typing import Optional, Any
import openpyxl
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
import csv
from io import StringIO, BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from app.services.utils.metrics import mape
from decimal import Decimal
from app.repositories.sales_repo import SalesRepository
from app.repositories.menu_items_repo import MenuItemRepository
from app.repositories.forecast_breakdown_repo import ForecastBreakdownRepository
from app.repositories.daily_forecast_accuracy_repo import DailyForecastAccuracyRepository
from app.repositories.restaurants_repo import RestaurantRepository
from app.repositories.activity_logs_repo import ActivityLogRepository
from app.utils.logger_helpers import log_method
from app.core.logging import logger 
from app.schemas.dashboard_dto import EodSalesEntriesIn
from typing import Sequence, Optional as Opt

class DashboardService:
    def __init__(self, db: AsyncSession, restaurant_id: int, subscription_tier: str, employee_id: int):
        self.db = db
        self.restaurant_id = restaurant_id
        self.subscription_tier = subscription_tier
        self.employee_id = employee_id
        self.sales_repo = SalesRepository(db, restaurant_id)
        self.menu_repo = MenuItemRepository(db, restaurant_id)
        self.restaurant_repo = RestaurantRepository(db,restaurant_id)
        self.forecast_breakdown_repo = ForecastBreakdownRepository(db, restaurant_id)
        self.daily_accuracy_repo = DailyForecastAccuracyRepository(db,restaurant_id)
        self.activity_log_repo = ActivityLogRepository(db,restaurant_id,employee_id)

    async def log_activity(self, action: str, details: Any = None):
        """
        Create an activity log entry safely.
        Converts details to JSON string if possible, else uses str().
        Adds timestamp to the log.
        """
        try:
            if details is None:
                details_str = ""
            else:
                try:
                    details_str = json.dumps(details, default=lambda o: o.__dict__, indent=2)
                except (TypeError, ValueError):
                    details_str = str(details)

            log_entry = {
                "action": action,
                "details": details_str,
                "created_at": datetime.utcnow().isoformat() + "Z"
            }

            await self.activity_log_repo.create(log_entry)
            logger.info(f"Activity logged: {action}")
        except Exception as e:
            logger.error(f"Failed to log activity '{action}': {e}", exc_info=True)

    async def get_daily_overview_data(self):
        if self.subscription_tier == 'basic':
            return await self._get_basic_overview()
        # Later: add 'plus', 'pro', etc.

    @log_method("Get basic overview")
    async def _get_basic_overview(self):
        return {
            "forecasted_sales_today": await self._get_forecasted_sales_basic(),
            "top_5_items_today": await self._get_top_forecasted_items_basic(),
            "accuracy_yesterday": await self._get_yesterday_accuracy_basic()
        }

    @log_method("Get Forecasted Sales (Basic)")
    async def _get_forecasted_sales_basic(self):
        today = date.today()
        forecast_rows = await self.forecast_breakdown_repo.get_forecasts_by_date(today)

        total_quantity = 0
        total_revenue = Decimal("0.00")

        for row in forecast_rows:
            menu_item = await self.menu_repo.get_by_id(row.menu_item_id)
            if menu_item:
                total_quantity += row.forecasted_quantity
                price = Decimal(str(menu_item.price))
                total_revenue += Decimal(row.forecasted_quantity) * price

        return {
            "forecasted_quantity": total_quantity,
            "forecasted_revenue": float(round(total_revenue, 2))  # convert Decimal to float for JSON serialization
        }

    @log_method("Get Top Forecasted Items (Basic)")
    async def _get_top_forecasted_items_basic(self):
        today = date.today()
        forecast_rows = await self.forecast_breakdown_repo.get_forecasts_by_date(today)

        item_quantities = defaultdict(int)

        for row in forecast_rows:
            item_quantities[row.menu_item_id] += row.forecasted_quantity

        # Sort by quantity descending
        sorted_items = sorted(item_quantities.items(), key=lambda x: x[1], reverse=True)

        results = []
        for menu_item_id, quantity in sorted_items:
            if len(results) >= 6:
                break

            item = await self.menu_repo.get_by_id(menu_item_id)
            if item and item.restaurant_id == self.restaurant_id and getattr(item, "is_active", True):
                results.append({
                    "menu_item_id": menu_item_id,
                    "name": item.name,
                    "forecasted_quantity": quantity,
                })

        return results
    
    @log_method("Get Yesterday Accuracy (Basic)")
    async def _get_yesterday_accuracy_basic(self):
        

        yesterday = date.today() - timedelta(days=1)
        accuracy_entries = await self.daily_accuracy_repo.get_by_date(yesterday)

        if not accuracy_entries:
            return {
                "accuracy_percent": None,
                "note": "No forecast accuracy data available"
            }

        mape_values = []
        for row in accuracy_entries:
            mape_value = mape(row.predicted_quantity, row.actual_quantity)
            mape_values.append(mape_value)

        avg_mape = sum(mape_values) / len(mape_values)
        accuracy_percent = round(max(0.0, 100 - avg_mape), 2)

        note = "Accurate" if accuracy_percent >= 85 else "Needs Improvement"

        return {
            "accuracy_percent": accuracy_percent,
            "note": note
        }

    @log_method("Create Menu Item")
    async def create_menu_item(self, data: dict):
        logger.info(f"Creating menu item: {data}")
        item = await self.menu_repo.create(data)
        await self.log_activity("create_menu_item", data)
        return item
    
    @log_method("Update Menu Item")
    async def update_menu_item(self, menu_item_id: int, data: dict):
        logger.info(f"Updating Menu Item: {menu_item_id}, {data}" )
        updated = await self.menu_repo.update(menu_item_id, data)
        details = {"menu_item_id": menu_item_id, **data}
        await self.log_activity("update_menu_item", details)
        return updated

    @log_method("Deactivate Menu Item")
    async def deactivate_menu_item(self, menu_item_id: int):
        logger.info("Deactivating menu item ID %d", menu_item_id)
        updated = await self.menu_repo.update(menu_item_id, {"is_active": False})
        await self.log_activity("deactivate_menu_item", updated)
        return updated
    
    @log_method("List Menu Items")
    async def list_menu_items(self):
        print('inside service')
        return await self.menu_repo.get_all()

    @log_method("Upload Bulk Menu Items")
    async def upload_menu_items_csv(self, file: UploadFile):
        contents = await file.read()
        created = []

        def normalize_headers(headers):
            return [header.strip().lower() if isinstance(header, str) else header for header in headers]

        if file.filename.endswith('.csv'):
            decoded = contents.decode('utf-8')
            raw_lines = decoded.splitlines()
            reader = csv.DictReader(raw_lines)
            reader.fieldnames = normalize_headers(reader.fieldnames)

            for row in reader:
                normalized_row = {k.strip().lower(): v for k, v in row.items()}
                try:
                    item_data = {
                        "name": normalized_row['name'],
                        "category": normalized_row.get('category'),
                        "price": float(normalized_row['price']),
                        "is_active": normalized_row.get('is_active', '1').lower() in ['1', 'true', 'yes']
                    }
                    item = await self.menu_repo.create(item_data)
                    created.append(item)
                    await self.log_activity("upload_menu_item", item_data)
                except Exception as e:
                    logger.error(f"Failed to process menu item row {row}: {e}", exc_info=True)
                    continue

        elif file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filename=BytesIO(contents), read_only=True)
            sheet = workbook.active

            raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            headers = normalize_headers(raw_headers)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                try:
                    item_data = {
                        "name": row_dict['name'],
                        "category": row_dict.get('category'),
                        "price": float(row_dict['price']),
                        "is_active": str(row_dict.get('is_active', '1')).lower() in ['1', 'true', 'yes']
                    }
                    item = await self.menu_repo.create(item_data)
                    created.append(item)
                    await self.log_activity("upload_menu_item",item_data)
                except Exception as e:
                    logger.error(f"Failed to process menu item row {row_dict}: {e}", exc_info=True)
                    continue

        else:
            raise ValueError("Unsupported file format. Please upload CSV or XLSX.")

        return created

    @log_method("Upload Sales Data")
    async def upload_sales_data(self, file: UploadFile, overwrite: bool = False):
        contents = await file.read()
        inserted_sales = []
        today = datetime.now()
    sale_dates = set()
    # Map date -> set of channels present in the upload (None for unspecified)
    date_channels = {}

        def normalize_headers(headers):
            return [header.strip().lower() if isinstance(header, str) else header for header in headers]

        parsed_rows = []

        if file.filename.endswith(".csv"):
            decoded = contents.decode("utf-8")
            raw_lines = decoded.splitlines()
            reader = csv.DictReader(raw_lines)
            reader.fieldnames = normalize_headers(reader.fieldnames)

            for row in reader:
                normalized_row = {k.strip().lower(): v for k, v in row.items()}
                parsed_rows.append(normalized_row)

        elif file.filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(filename=BytesIO(contents), read_only=True)
            sheet = workbook.active
            headers = normalize_headers([cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))])

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                parsed_rows.append(row_dict)

        else:
            raise ValueError("Unsupported file format. Please upload CSV or XLSX.")

        # Collect sale dates and per-date channels
        for row in parsed_rows:
            timestamp = row.get("sale_timestamp")
            try:
                sale_time = datetime.fromisoformat(timestamp) if timestamp else today
                sale_dates.add(sale_time.date())
                ch_val = row.get("sales_channel")
                ch = ch_val if (ch_val is not None and ch_val != "") else None
                d = sale_time.date()
                s = date_channels.get(d)
                if s is None:
                    s = set()
                    date_channels[d] = s
                s.add(ch)
            except Exception as e:
                logger.error(f"Skipping row with invalid timestamp '{timestamp}': {e}", exc_info=True)

        # Check conflicts per date+channel
        if not overwrite:
            conflict = False
            for d, channels in date_channels.items():
                if await self.sales_repo.sales_exist_for_date_and_channels(d, list(channels)):
                    conflict = True
                    break
            if conflict:
                raise HTTPException(
                    status_code=409,
                    detail="Sales already exist for one or more date/channel combinations in this file. Confirm overwrite to replace those channels."
                )
        else:
            # Delete only the channels present in the file for each date
            for d, channels in date_channels.items():
                await self.sales_repo.delete_sales_for_date_and_channels(d, list(channels))
            await self.log_activity("overwrite_sales_file", {"dates": [d.isoformat() for d in sale_dates], "channels": {k.isoformat(): list(v) for k, v in date_channels.items()}})

        # Process each row
        for row in parsed_rows:
            try:
                menu_item_id = None
                if "menu_item_id" in row and row["menu_item_id"]:
                    menu_item_id = int(row["menu_item_id"])
                elif "menu_item_name" in row and row["menu_item_name"]:
                    item = await self.menu_repo.get_by_name(row["menu_item_name"])
                    if not item:
                        logger.error(f"Menu item not found: {row['menu_item_name']}")
                        continue
                    menu_item_id = item.id

                if not menu_item_id:
                    logger.error("Skipping row with missing menu_item_id")
                    continue

                quantity = int(row["quantity_sold"])
                timestamp = row.get("sale_timestamp")
                sale_time = datetime.fromisoformat(timestamp) if timestamp else today
                sales_channel = row.get("sales_channel")

                sale_data = {
                    "restaurant_id": self.restaurant_id,
                    "menu_item_id": menu_item_id,
                    "quantity_sold": quantity,
                    "sales_channel": sales_channel,
                    "sale_timestamp": sale_time
                }

                sale = await self.sales_repo.create(sale_data)
                inserted_sales.append(sale)
                # await self.log_activity("upload_sales_data",sale_data)

            except Exception as e:
                logger.error(f"Failed to process sales row {row}: {e}", exc_info=True)
                continue

        return inserted_sales

    @log_method("Upload Sales Entries (Manual JSON)")
    async def upload_sales_entries(self, payload: EodSalesEntriesIn):
        """Insert manual EOD sales entries for a single date."""
        try:
            sale_date = datetime.fromisoformat(payload.sale_date).date()
        except Exception:
            raise ValueError("Invalid sale_date; expected YYYY-MM-DD")

        # Determine channels present in this submission
        channels = []
        try:
            channels = list({ (e.sales_channel if getattr(e, 'sales_channel', None) is not None else None) for e in payload.entries })
        except Exception:
            channels = [None]

        # Overwrite semantics by channel (not entire date)
        if payload.overwrite:
            await self.sales_repo.delete_sales_for_date_and_channels(sale_date, channels)
            await self.log_activity("overwrite_sales_manual", {"sale_date": payload.sale_date, "channels": channels})
        else:
            # if existing for any of the channels and not overwrite, raise 409 like file upload
            if await self.sales_repo.sales_exist_for_date_and_channels(sale_date, channels):
                raise HTTPException(status_code=409, detail="Sales already exist for that date and channel(s). Confirm overwrite to replace.")

        inserted = []
        for entry in payload.entries:
            try:
                # Ensure menu item exists and belongs to restaurant
                item = await self.menu_repo.get_by_id(entry.menu_item_id)
                if not item or item.restaurant_id != self.restaurant_id:
                    logger.error(f"Menu item not found or unauthorized: {entry.menu_item_id}")
                    continue

                sale_time = datetime.combine(sale_date, datetime.min.time())
                sale_data = {
                    "restaurant_id": self.restaurant_id,
                    "menu_item_id": entry.menu_item_id,
                    "quantity_sold": int(entry.quantity_sold),
                    "sales_channel": entry.sales_channel,
                    "sale_timestamp": sale_time,
                }
                s = await self.sales_repo.create(sale_data)
                inserted.append(s)
            except Exception as e:
                logger.error(f"Failed to insert manual sale entry {entry}: {e}", exc_info=True)
                continue

        return inserted
    
    @log_method("Generating Sales Upload Template")
    async def generate_sales_upload_template_xlsx(self, default_date: Optional[str] = None) -> BytesIO:
        menu_items = await self.list_menu_items()
        active_menu_items = [item for item in menu_items if item.is_active]
        restaurant_settings = await self.restaurant_repo.get_settings()
        sales_channels = restaurant_settings['sales_channels']

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales Upload Template"

        headers = ["menu_item_id", "menu_item_name", "quantity_sold", "sales_channel", "sale_timestamp"]
        sheet.append(headers)

        if default_date is None:
            default_date = datetime.utcnow().strftime("%Y-%m-%dT00:00:00")

        # Append all rows first
        for item in active_menu_items:
            for channel in sales_channels:
                sheet.append([item.menu_item_id, item.name, "", channel, default_date])

        total_rows = sheet.max_row
        total_cols = len(headers)

        # Lock all cells by default
        for row in range(1, total_rows + 1):
            for col in range(1, total_cols + 1):
                sheet.cell(row=row, column=col).protection = Protection(locked=True)

        # Unlock quantity_sold column cells (column 3), skipping header
        for row in range(2, total_rows + 1):
            sheet.cell(row=row, column=3).protection = Protection(locked=False)

        # Set column widths based on max content length per column
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(header)
            col_letter = get_column_letter(col_idx)
            for row in range(1, total_rows + 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[col_letter].width = max_length + 2

        # Protect the sheet
        sheet.protection.sheet = True
        sheet.protection.password = "restaurant"
        sheet.protection.enable()

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @log_method("Check sales conflicts by date/channels")
    async def check_sales_conflicts(self, sale_date: str, channels: Opt[Sequence[str]] = None):
        """Return per-channel counts for a date, to support client confirmation UX.
        channels: if provided, filter to those channels; use the string 'null' to represent None.
        """
        try:
            d = datetime.fromisoformat(sale_date).date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid sale_date; expected YYYY-MM-DD")

        # If channels not provided, return counts for all channels present
        if not channels:
            rows = await self.sales_repo.get_sales_channels_counts_for_date(d)
            return {
                "sale_date": sale_date,
                "conflicts": { (ch if ch is not None else None): cnt for ch, cnt in rows },
            }
        # Normalize 'null' -> None
        norm_channels = [None if ch.lower() == 'null' else ch for ch in channels]
        exists = await self.sales_repo.sales_exist_for_date_and_channels(d, norm_channels)
        if not exists:
            return {"sale_date": sale_date, "conflicts": {}}
        # Return counts but filtered to requested channels
        rows = await self.sales_repo.get_sales_channels_counts_for_date(d)
        filtered = { (ch if ch is not None else None): cnt for ch, cnt in rows if (ch in norm_channels) or (ch is None and None in norm_channels) }
        return {"sale_date": sale_date, "conflicts": filtered}

    